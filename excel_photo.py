import os
import sys
from tkinter import Image
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

#first argment for excel file name
excel_file=input("file name: ", )  #run: python excel_photo.py test.xls
if(os.path.isfile(excel_file)):
    print("working on file:", excel_file)
else:
    print("file not exist")
    quit()

""" password check for tweety only
passwd = input("Password: ", )
if(passwd != "248622"):
    print("bye!")
    quit()
"""


workbook = openpyxl.load_workbook(excel_file)

worksheetname = input("worksheet name:")

#worksheetname = '工作表1' #hardcode for debugging
worksheet = workbook[worksheetname]

print("working on worksheet:", worksheetname)
#print(worksheet['A1'].value,worksheet['B1'].value)
range_row = worksheet.max_row
print("worksheet:",worksheetname,"has",range_row-1, "rows for image insert")
path = os.getcwd() + '/'
print("current path:",os.getcwd())
for row in range(4,worksheet.max_row):
    jpgname=str(worksheet['B'+str(row)].value)
    print("for row number",row ," looking for JPG start with", jpgname)
    #searh jpg start with jpgname by case insensitive
    #if row >373:
    #    break
    for f in os.listdir(str(path)):
        upperf = f.upper()
        #print(upperf)
        if os.path.isfile(path + upperf):
            if upperf.startswith(jpgname):#jpg found
                #filepath = str(os.getcwd())+"\\" +str(f)
                print(jpgname, "found: ",f)
                img = Image(str(f))
                oldwidth = img.width
                oldheight =img.height
                if img.width>img.height:
                    img.width = worksheet.column_dimensions['A'].width
                    img.height = oldheight*img.width/oldwidth
                else:
                    img.height = worksheet.row_dimensions[row].height
                    img.width = oldwidth*img.height/oldheight
                print("Adding jpg",f," to cell A",row)
                worksheet.add_image(img,'A'+ str(row))          
        #else:
            #print(jpgname,"not found")
newfilename = input("please enter new file name(end with .xlsx):")    
workbook.save(str(newfilename))            
